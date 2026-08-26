import os
import json
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, LineChart, Reference

def list_json_files():
    files = [f for f in os.listdir('.') if f.startswith("Large_Oven_Status_") and f.endswith(".json")]
    files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
    return files

def paginate_files(files):
    page = 0
    while True:
        start, end = page * 10, page * 10 + 10
        subset = files[start:end]
        if not subset:
            print("No more files.")
            return None
        print("\nSelect a file to process:")
        for i, f in enumerate(subset, 1):
            print(f"{i}. {f}")
        choice = input("Enter 1–10 to select, 'c' for next page, or 'q' to quit: ").strip().lower()
        if choice == 'q':
            return None
        elif choice == 'c':
            page += 1
        elif choice.isdigit() and 1 <= int(choice) <= len(subset):
            return subset[int(choice) - 1]
        else:
            print("Invalid input. Try again.")

def flatten_record(rec):
    flat = {}
    flat["timestamp"] = rec.get("timestamp")
    # Flatten nested dictionaries
    for group_key in ("status", "indicators", "safety_status"):
        group = rec.get(group_key, {})
        for k, v in group.items():
            flat[f"{group_key}_{k}"] = v
    return flat

def load_json_data(filename):
    with open(filename, 'r') as f:
        data = json.load(f)
    # Flatten all records
    return [flatten_record(r) for r in data]

def calculate_utilization(df):
    """Assume 'in use' means status_state not in ['FAULT', 'IDLE']."""
    df["timestamp"] = pd.to_datetime(df["timestamp"])
    df["date"] = df["timestamp"].dt.date
    df["in_use"] = ~df["status_state"].isin(["FAULT", "IDLE"])
    
    # Each row = 1 second sample assumed
    daily = df.groupby("date")["in_use"].mean().reset_index()
    daily["utilization_percent"] = daily["in_use"] * 100
    return daily

def save_to_excel(df, daily_util, filename):
    xlsx_name = filename.replace(".json", ".xlsx")

    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    # --- RAW DATA SHEET ---
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_raw.append(r)

    # --- PIE CHART SHEET ---
    ws_pie = wb.create_sheet("Status Breakdown")
    status_counts = df["status_state"].value_counts().reset_index()
    status_counts.columns = ["State", "Count"]

    for r in dataframe_to_rows(status_counts, index=False, header=True):
        ws_pie.append(r)

    pie = PieChart()
    pie.title = "Status Breakdown"
    labels = Reference(ws_pie, min_col=1, min_row=2, max_row=len(status_counts)+1)
    data_vals = Reference(ws_pie, min_col=2, min_row=1, max_row=len(status_counts)+1)
    pie.add_data(data_vals, titles_from_data=True)
    pie.set_categories(labels)
    ws_pie.add_chart(pie, "D2")

    # --- UTILIZATION CHART SHEET ---
    ws_line = wb.create_sheet("Daily Utilization")
    for r in dataframe_to_rows(daily_util, index=False, header=True):
        ws_line.append(r)

    line_chart = LineChart()
    line_chart.title = "Utilization per Day"
    line_chart.x_axis.title = "Date"
    line_chart.y_axis.title = "Utilization (%)"
    data_ref = Reference(ws_line, min_col=3, min_row=1, max_row=len(daily_util)+1)
    cats_ref = Reference(ws_line, min_col=1, min_row=2, max_row=len(daily_util)+1)
    line_chart.add_data(data_ref, titles_from_data=True)
    line_chart.set_categories(cats_ref)
    ws_line.add_chart(line_chart, "E2")

    wb.save(xlsx_name)
    print(f"\n✅ Saved Excel file as {xlsx_name}")

def main():
    files = list_json_files()
    if not files:
        print("No matching JSON files found.")
        return

    selected = paginate_files(files)
    if not selected:
        print("No file selected. Exiting.")
        return

    data = load_json_data(selected)
    if not data:
        print("File is empty or invalid.")
        return

    df = pd.DataFrame(data)
    daily_util = calculate_utilization(df)
    save_to_excel(df, daily_util, selected)

if __name__ == "__main__":
    main()
